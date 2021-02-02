import re
import datetime
import openpyxl
from io import BytesIO
import urllib.request
from .statics import inhabitants

def getFile():
  # Request to load excel sheet
  url = 'https://www.rki.de/DE/Content/InfAZ/N/Neuartiges_Coronavirus/Daten/Impfquotenmonitoring.xlsx?__blob=publicationFile'
  hdr = {
      'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11'}

  req = urllib.request.Request(url, headers=hdr)
  response = urllib.request.urlopen(req)

  return response.read()

def getData():
  states = inhabitants.states

  file = getFile()

  # Read excel sheet
  wb = openpyxl.load_workbook(filename = BytesIO(file)) 
  sheet = wb[wb.sheetnames[1]]

  # Load update time
  lastUpdateRawString = wb.sheetnames[1]
  relastUpdateMatch = re.search(r"[\d]{2}\.[\d]{2}\.[\d]{2}", lastUpdateRawString)
  lastUpdate = datetime.datetime.strptime(relastUpdateMatch.group(), '%d.%m.%y')

  # Load data from rows
  sumStates = 0
  sumStates2nd = 0
  sumDiffStates = 0
  sumDiffStates2nd = 0
  for row in sheet.iter_rows(max_row=19):
    if row[1].value is None:
      continue
    state = row[1].value.replace("*", "")
    if state in states:
      states[state]['rs'] = str(row[0].value)

      # First vaccination
      states[state]['vaccinated'] = row[3].value
      states[state]['vaccinated_by_accine'] = {}
      states[state]['vaccinated_by_accine']['biontech'] = row[4].value
      states[state]['vaccinated_by_accine']['moderna'] = row[5].value
      states[state]['difference_to_the_previous_day'] = row[6].value
      states[state]['vaccinations_per_1000_inhabitants'] = round(states[state]['vaccinated'] / states[state]['total'] * 1000, 2)
      states[state]['quote'] = round(row[7].value, 2)

      sumStates += states[state]['vaccinated']
      sumDiffStates += states[state]['difference_to_the_previous_day']


      # Second vaccination
      states[state]['2nd_vaccination'] = {}
      states[state]['2nd_vaccination']['vaccinated'] = row[8].value
      states[state]['2nd_vaccination']['difference_to_the_previous_day'] = row[9].value

      sumStates2nd += states[state]['2nd_vaccination']['vaccinated']
      
      if states[state]['2nd_vaccination']['difference_to_the_previous_day'] is not None:
        sumDiffStates2nd += states[state]['2nd_vaccination']['difference_to_the_previous_day']
  
  return {
    "lastUpdate": lastUpdate,
    "states": states,
    "sumStates": sumStates,
    "sumStates2nd": sumStates2nd,
    "sumDiffStates": sumDiffStates,
    "sumDiffStates2nd": sumDiffStates2nd,
    "totalGermany": inhabitants.total
  }
